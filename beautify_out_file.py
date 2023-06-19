import openpyxl
from openpyxl.styles import Font, Fill, PatternFill
from config import PATH


def beautify_grid_maks(file):
    # print(3, file)
    wb = openpyxl.load_workbook(PATH + file)
    # print(type(wb))
    # print(wb.sheetnames)
    sheet = wb.active
    # print(sheet['A2'].value)
    rows, cols = sheet.max_row, sheet.max_column
    # print(rows, cols)
    # print(3, sheet.cell(row=rows, column=cols).value)
    # sheet.column_dimensions['B'] = 100
    row_count = 1
    while row_count != rows + 1:
        if sheet.cell(row=row_count, column=cols).value == 'Выдан на руки':
            sheet.cell(row=row_count, column=cols - 2).fill = PatternFill("solid", fgColor="00FF00")


            if sheet.cell(row=row_count, column=cols - 6).value[1:] == sheet.cell(row=row_count, column=cols - 2).value:
                sheet.cell(row=row_count, column=cols - 6).fill = PatternFill("solid", fgColor="00FF00")
                # sheet.cell(row=row_count, column=cols + 1).value = 'Полис совпадает с полисом в ТФОМС'
            else:
                sheet.cell(row=row_count, column=cols - 6).fill = PatternFill("solid", fgColor="FF0000")
                # sheet.cell(row=row_count, column=cols + 1).value = 'Полис не совпадает с полисом в ТФОМС'

        elif sheet.cell(row=row_count, column=cols).value == 'Не найден, уточните персональные данные':
            sheet.cell(row=row_count, column=cols - 3).fill = PatternFill("solid", fgColor="FF0000")
            sheet.cell(row=row_count, column=cols - 8).fill = PatternFill("solid", fgColor="FF0000")
            # sheet.cell(row=row_count, column=cols + 1).value = 'Пациент не найден в базе ТФОМС'

        elif sheet.cell(row=row_count, column=cols).value == 'Снят с учета':
            for col_ in range(9):
                sheet.cell(row=row_count, column=cols - col_).fill = PatternFill("solid", fgColor="FF0000")
                # sheet.cell(row=row_count, column=cols + 1).value = 'Пациент снят с учета'

        elif sheet.cell(row=row_count, column=cols).value[:9] == 'Отпечатан':
            for col_ in range(9):
                sheet.cell(row=row_count, column=cols - col_).fill = PatternFill("solid", fgColor="FFFF00")
                # sheet.cell(row=row_count, column=cols + 1).value = 'Пациент получил временный полис'
            sheet.cell(row=row_count, column=cols - 2).fill = PatternFill("solid", fgColor="00FF00")

        elif sheet.cell(row=row_count, column=cols).value[:36] == 'Снят c учета по причине дублирования':
            sheet.cell(row=row_count, column=cols - 2).value = sheet.cell(row=row_count, column=cols).value[-16:]
            sheet.cell(row=row_count, column=cols - 2).fill = PatternFill("solid", fgColor="00FF00")
            # sheet.cell(row=row_count, column=cols + 1).value = 'Попробуйте занести указанный полис'

        row_count += 1

    # print(file)
    # send_document_handler(file)
    wb.save(PATH + file)


def beautify_grid_rmis(file):
    wb = openpyxl.load_workbook(PATH + file)
    sheet = wb.active
    rows, cols = sheet.max_row, sheet.max_column

    row_count = 1
    while row_count != rows + 1:
        if sheet.cell(row=row_count, column=cols).value == 'Выдан на руки':
            sheet.cell(row=row_count, column=cols - 2).fill = PatternFill("solid", fgColor="00FF00")

        elif sheet.cell(row=row_count, column=cols).value == 'Не найден, уточните персональные данные':
            sheet.cell(row=row_count, column=cols - 4).fill = PatternFill("solid", fgColor="FF0000")
            sheet.cell(row=row_count, column=cols - 5).fill = PatternFill("solid", fgColor="FF0000")
            # sheet.cell(row=row_count, column=cols + 1).value = 'Пациент не найден в базе ТФОМС'

        elif sheet.cell(row=row_count, column=cols).value == 'Снят с учета':
            for col_ in range(6):
                sheet.cell(row=row_count, column=cols - col_).fill = PatternFill("solid", fgColor="FF0000")
                # sheet.cell(row=row_count, column=cols + 1).value = 'Пациент снят с учета'

        elif sheet.cell(row=row_count, column=cols).value[:9] == 'Отпечатан':
            for col_ in range(6):
                sheet.cell(row=row_count, column=cols - col_).fill = PatternFill("solid", fgColor="FFFF00")
                # sheet.cell(row=row_count, column=cols + 1).value = 'Пациент получил временный полис'
            sheet.cell(row=row_count, column=cols - 2).fill = PatternFill("solid", fgColor="00FF00")

        elif sheet.cell(row=row_count, column=cols).value[:36] == 'Снят c учета по причине дублирования':
            sheet.cell(row=row_count, column=cols - 2).value = sheet.cell(row=row_count, column=cols).value[-16:]
            sheet.cell(row=row_count, column=cols - 2).fill = PatternFill("solid", fgColor="00FF00")
            # sheet.cell(row=row_count, column=cols + 1).value = 'Попробуйте занести указанный полис'

        row_count += 1

    # print(file)
    # send_document_handler(file)
    wb.save(PATH + file)


def beautify_grid_foms(file):
    wb = openpyxl.load_workbook(PATH + file)
    sheet = wb.active
    rows, cols = sheet.max_row, sheet.max_column
    print(rows,cols)
    #  Приведем в порядок столбцы в excel

    sheet.column_dimensions['A'].width = 48
    sheet.column_dimensions['B'].width = 18
    sheet.column_dimensions['C'].width = 40
    sheet.column_dimensions['D'].width = 14
    sheet.column_dimensions['E'].width = 32
    sheet.column_dimensions['F'].width = 18
    sheet.column_dimensions['G'].width = 35
    sheet.column_dimensions['H'].width = 35
    sheet.column_dimensions['I'].width = 25
    sheet.column_dimensions['J'].width = 13
    sheet.column_dimensions['K'].width = 18
    sheet.column_dimensions['L'].width = 14
    sheet.column_dimensions['M'].width = 14

    sheet.cell(1, 6).value = 'ЕНП'
    sheet.cell(1, 7).value = 'Статус'
    sheet.cell(1, 8).value = 'СМО'
    sheet.cell(1, 9).value = 'Регион'
    sheet.cell(1, 10).value = 'Серия полиса'
    sheet.cell(1, 11).value = 'Номер полиса'
    sheet.cell(1, 12).value = 'Дата получения'
    sheet.cell(1, 13).value = 'Дата прекращения'

    row_count = 2
    while row_count != rows + 1:
        # print(rows, cols, row_count)
        if sheet.cell(row_count, 7).value == "Снят с учёта в Республике Дагестан":
            sheet.cell(row_count, 7).fill = PatternFill("solid", fgColor="0000FF00")
            sheet.cell(row_count, 8).fill = PatternFill("solid", fgColor="0000FF00")
            sheet.cell(row_count, 9).fill = PatternFill("solid", fgColor="0000FF00")

        if sheet.cell(row_count, 13).value != None:
            sheet.cell(row_count, 7).fill = PatternFill("solid", fgColor="00FF0000")
            sheet.cell(row_count, 8).fill = PatternFill("solid", fgColor="00FF0000")
            sheet.cell(row_count, 9).fill = PatternFill("solid", fgColor="00FF0000")
            sheet.cell(row_count, 13).fill = PatternFill("solid", fgColor="00FF0000")

        if sheet.cell(row_count, 9).value == 'Дагестан' and sheet.cell(row_count, 7).value == 'Действующий':
            sheet.cell(row_count, 7).fill = PatternFill("solid", fgColor="00FFFF00")
            sheet.cell(row_count, 8).fill = PatternFill("solid", fgColor="00FFFF00")
            sheet.cell(row_count, 9).fill = PatternFill("solid", fgColor="00FFFF00")
            sheet.cell(row_count, 12).fill = PatternFill("solid", fgColor="00FFFF00")

        if sheet.cell(row_count, 9).value == 'Дагестан' and sheet.cell(row_count, 7).value ==\
                'Снят с учёта в Республике Дагестан':
            sheet.cell(row_count, 7).fill = PatternFill("solid", fgColor="00FF0000")
            sheet.cell(row_count, 8).fill = PatternFill("solid", fgColor="00FF0000")
            sheet.cell(row_count, 9).fill = PatternFill("solid", fgColor="00FF0000")

        if sheet.cell(row_count, 7).value == 'Нет такого в реестре!':
            sheet.cell(row_count, 3).fill = PatternFill("solid", fgColor="00FF0000")
            sheet.cell(row_count, 4).fill = PatternFill("solid", fgColor="00FF0000")
            sheet.cell(row_count, 7).fill = PatternFill("solid", fgColor="00FF0000")

        print(sheet.cell(row_count, 11).value, sheet.cell(row_count, 12).value)
        if sheet.cell(row_count, 11).value == None and sheet.cell(row_count, 12).value == None:
            sheet.cell(row_count, 7).fill = PatternFill("solid", fgColor="00FF0000")
            sheet.cell(row_count, 11).fill = PatternFill("solid", fgColor="00FF0000")
            sheet.cell(row_count, 12).fill = PatternFill("solid", fgColor="00FF0000")


        row_count += 1




        # if sheet.cell(row=row_count, column=cols).value == 'Выдан на руки':
        #     sheet.cell(row=row_count, column=cols - 2).fill = PatternFill("solid", fgColor="00FF00")
        #
        # elif sheet.cell(row=row_count, column=cols).value == 'Не найден, уточните персональные данные':
        #     sheet.cell(row=row_count, column=cols - 4).fill = PatternFill("solid", fgColor="FF0000")
        #     sheet.cell(row=row_count, column=cols - 5).fill = PatternFill("solid", fgColor="FF0000")
        #
        # elif sheet.cell(row=row_count, column=cols).value == 'Снят с учета':
        #     for col_ in range(6):
        #         sheet.cell(row=row_count, column=cols - col_).fill = PatternFill("solid", fgColor="FF0000")
        #         # sheet.cell(row=row_count, column=cols + 1).value = 'Пациент снят с учета'
        #
        # elif sheet.cell(row=row_count, column=cols).value[:9] == 'Отпечатан':
        #     for col_ in range(6):
        #         sheet.cell(row=row_count, column=cols - col_).fill = PatternFill("solid", fgColor="FFFF00")
        #         # sheet.cell(row=row_count, column=cols + 1).value = 'Пациент получил временный полис'
        #     sheet.cell(row=row_count, column=cols - 2).fill = PatternFill("solid", fgColor="00FF00")
        #
        # elif sheet.cell(row=row_count, column=cols).value[:36] == 'Снят c учета по причине дублирования':
        #     sheet.cell(row=row_count, column=cols - 2).value = sheet.cell(row=row_count, column=cols).value[-16:]
        #     sheet.cell(row=row_count, column=cols - 2).fill = PatternFill("solid", fgColor="00FF00")
            # sheet.cell(row=row_count, column=cols + 1).value = 'Попробуйте занести указанный полис'



    # print(file)
    # send_document_handler(file)
    wb.save(PATH + file)