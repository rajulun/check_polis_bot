import openpyxl
import xlrd
import datetime
import pandas as pd
import numpy as np
from utils_for_request import get_polis_number, get_begin_row_maks_excel
from beautify_out_file import beautify_grid_maks, beautify_grid_rmis, beautify_grid_foms
from config import PATH


def get_polis_from_maks_exel(file):
    df = pd.read_excel(io=file, engine='openpyxl')
    begin_row_maks_excel = get_begin_row_maks_excel(df)
    print(begin_row_maks_excel)
    df = pd.read_excel(io=file,
                       engine='openpyxl',
                       usecols='C:H',
                       header=begin_row_maks_excel+2,  # в excel это №5
                       )
    df.drop([0], inplace=True)
    print(df)
    col_names = df.columns
    print('cloumns: ' + col_names)
    # df = df.drop(columns={col_names[0], col_names[1]})
    print(df)
    df = df.drop_duplicates()
    df = df.loc[(df[col_names[1]] == 'НЕ НАЙДЕН В БД СРЗ')]
    df = df.assign(foms_num=np.nan, foms_smo=np.nan, foms_comment=np.nan)
    df.reset_index(inplace=True, drop=True)
    print(df)
    i = begin_row_maks_excel
    for i in range(df.shape[0]):
        fio = df[col_names[0]][i]
        dr = df[col_names[5]][i]
        df[col_names[5]][i] = np.where(type(dr) == datetime.datetime, dr.date(), dr.strftime("%d.%m.%Y"))

        if len(fio.split()) == 3:
            fam, im, ot = fio.split()
        elif len(fio.split()) == 2:
            fam, im = fio.split()
        else:
            fam, im, ot, temp_ = fio.split()
            ot = ot + ' ' + temp_
        dayr = dr.strftime("%d.%m.%Y")
        # dayr1 = dr
        # print(fam, im, ot, dayr)
        req_polis_number = get_polis_number(fam, im, ot, dayr)
        # print(req_polis_number)
        if len(req_polis_number) > 1:
            df['foms_num'][i] = str(req_polis_number[0])
            df['foms_smo'][i] = req_polis_number[1]
            df['foms_comment'][i] = req_polis_number[2]

        else:
            df['foms_num'][i] = "Не найден, уточните персональные данные"
            df['foms_smo'][i] = "Не найден, уточните персональные данные"
            df['foms_comment'][i] = "Не найден, уточните персональные данные"


    print(df.shape[0])
    checked_file = 'checked_' + file
    # print(checked_file)

    df.to_excel(PATH + checked_file)
    beautify_grid_maks(checked_file)

def get_polis_from_rmis_exel(file):
    df = pd.read_excel(io=file,
                       engine='openpyxl',
                       usecols='A:C',)
    df.drop([0], inplace=True)
    print(df)

    col_names = df.columns
    print('columns: ' + col_names)
    # df = df.drop(columns={col_names[0], col_names[1]})
    df = df.drop_duplicates()
    df = df.loc[(df[col_names[2]].str.find('действительный полис ОМС')) > 0]
    # df = df.loc[(df[col_names[2]] == 'Отсутствует действительный полис ОМС')]
    df = df.assign(foms_num=np.nan, foms_smo=np.nan, foms_comment=np.nan)
    df.reset_index(inplace=True, drop=True)

    i = 1
    for i in range(df.shape[0]):
        fio = df[col_names[0]][i]
        dr = df[col_names[1]][i]
        # df[col_names[1]][i] = np.where(type(dr) == datetime.datetime, dr.date(), dr.strftime("%d.%m.%Y"))

        if len(fio.split()) == 3:
            fam, im, ot = fio.split()
        elif len(fio.split()) == 2:
            fam, im = fio.split()
        else:
            fam, im, ot, temp_ = fio.split()
            ot = ot + ' ' + temp_

        dayr = dr.split(' ')[0]
        print(fam, im, ot, dayr)
        req_polis_number = get_polis_number(fam, im, ot, dayr)
        if len(req_polis_number) > 1:
            df['foms_num'][i] = str(req_polis_number[0])
            df['foms_smo'][i] = req_polis_number[1]
            df['foms_comment'][i] = req_polis_number[2]

        else:
            df['foms_num'][i] = "Не найден, уточните персональные данные"
            df['foms_smo'][i] = "Не найден, уточните персональные данные"
            df['foms_comment'][i] = "Не найден, уточните персональные данные"

    checked_file = 'checked_' + file
    df.to_excel(PATH + checked_file)
    beautify_grid_rmis(checked_file)



def get_polis_from_foms_exel(file):
    df = pd.read_excel(io=file,
                       engine='openpyxl',
                       usecols='A:E',
                       )

    col_names = df.columns
    print('columns: ' + col_names)

    df = df.loc[(df[col_names[0]].str.find('Полис не найден в БД, либо недействительный')) == 0]
    df = df.drop_duplicates(subset=['Полис'])
    print(2, df.shape)

    df = df.assign(foms_enp=np.nan, foms_status=np.nan, foms_smo=np.nan,\
                   foms_region=np.nan, foms_spolis=np.nan, foms_npolis=np.nan,\
                   foms_startDate=np.nan, foms_endDate=np.nan)
    df.reset_index(inplace=True, drop=True)

    i = 1
    for i in range(df.shape[0]):
        print(df[col_names[2]][i])
        fio = df[col_names[2]][i]
        dr = df[col_names[3]][i]


        if len(fio.split()) == 3:
            fam, im, ot = fio.split()
        elif len(fio.split()) == 2:
            fam, im = fio.split()
        else:
            fam, im, ot, temp_ = fio.split()
            ot = ot + ' ' + temp_

        dayr = dr #.split(' ')[0]
        # print(1, fam, im, ot, dayr)
        req_polis_number = get_polis_number(fam, im, ot, dayr)

        if req_polis_number['result'] < 1:
            if req_polis_number['insuranceStatuses']['enp'] != None:
                df['foms_enp'][i] = '*' + req_polis_number['insuranceStatuses']['enp']
            df['foms_status'][i] = req_polis_number['insuranceStatuses']['status']
            df['foms_smo'][i] = req_polis_number['insuranceStatuses']['smo']
            df['foms_region'][i] = req_polis_number['insuranceStatuses']['region']
            df['foms_spolis'][i] = req_polis_number['insuranceStatuses']['spolis']
            if req_polis_number['insuranceStatuses']['npolis'] != None:
                df['foms_npolis'][i] = '*' + req_polis_number['insuranceStatuses']['npolis']
            df['foms_startDate'][i] = req_polis_number['insuranceStatuses']['startDate']
            df['foms_endDate'][i] = req_polis_number['insuranceStatuses']['endDate']

        else:
            df['foms_status'][i] = 'Нет такого в реестре!'

    checked_file = 'checked_' + file
    df.to_excel(PATH + checked_file, index=False)
    print(df)
    beautify_grid_foms(checked_file)


def used_users(id, username, name_i, name_f, text, sended_file, now, get_last_row_column):
    file = PATH + 'users_file.xlsx'
    wb = openpyxl.load_workbook(file)
    sheet = wb.active
    rows, cols = sheet.max_row, sheet.max_column
    sheet.cell(row=rows + 1, column=1).value = id
    sheet.cell(row=rows + 1, column=2).value = username
    sheet.cell(row=rows + 1, column=3).value = name_i
    sheet.cell(row=rows + 1, column=4).value = name_f
    sheet.cell(row=rows + 1, column=5).value = text
    sheet.cell(row=rows + 1, column=6).value = sended_file
    sheet.cell(row=rows + 1, column=7).value = now.strftime("%d-%m-%Y %H:%M")
    sheet.cell(row=rows + 1, column=8).value = get_last_row_column
    print(sheet.max_row)
    wb.save(file)


def get_last_row_column(file):
    wb = openpyxl.load_workbook(file)
    sheet = wb.active
    rows, cols = sheet.max_row, sheet.max_column
    return rows, cols